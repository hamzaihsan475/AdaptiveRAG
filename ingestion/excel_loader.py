"""
excel_loader.py

Loader for extracting text content from Excel (.xlsx) files
using openpyxl, converting each row into a readable text line.
"""

from openpyxl import load_workbook
from ingestion.base_loader import BaseLoader


class ExcelLoader(BaseLoader):
    """
    Loads and extracts text from Excel spreadsheets, converting
    each row into a single line of comma-separated text.
    """

    def load(self, file_path: str) -> str:
        """
        Extract text from all sheets of an Excel workbook.

        Args:
            file_path (str): Path to the .xlsx file.

        Returns:
            str: Text representation of all rows across all sheets.
        """
        workbook = load_workbook(file_path, data_only=True)
        lines = []
        for sheet in workbook.worksheets:
            lines.append(f"--- Sheet: {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                row_text = ", ".join(str(cell) for cell in row if cell is not None)
                if row_text:
                    lines.append(row_text)
        return "\n".join(lines)